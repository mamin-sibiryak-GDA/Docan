import os
import pathlib
import sys

if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    os.environ["PYMORPHY2_DICT_PATH"] = str(pathlib.Path(sys._MEIPASS).joinpath('pymorphy2_dicts_ru/data'))
from natasha import (
    Segmenter,
    MorphVocab,

    NewsEmbedding,
    NewsMorphTagger,
    NewsSyntaxParser,
    NewsNERTagger,

    Doc
)
import fitz
import pytesseract
import cv2
from cv2 import dnn_superres
from tqdm import tqdm
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
import numpy as np
import re

sr = dnn_superres.DnnSuperResImpl_create()
sr.readModel("./src/FSRCNN_x2.pb")
sr.setModel("fsrcnn", 2)


def get_page_text(current_page, doc):
    if doc.load_page(current_page.number).get_text("text") == "":
        for img in tqdm(doc.get_page_images(current_page.number),
                        desc="Обрабатывается %i страница из %i" % (current_page.number + 1, doc.page_count)):
            xref = img[0]
            image = doc.extract_image(xref)
            pix = fitz.Pixmap(doc, xref)

        image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
        image_upscale = sr.upsample(image)

        page_text = pytesseract.image_to_string(image_upscale, lang='rus')
        return page_text
    else:
        page = doc.load_page(current_page.number)
        page_text = page.get_text("text")
        return page_text


def doctype1():
    segmenter = Segmenter()
    morph_vocab = MorphVocab()

    emb = NewsEmbedding()
    morph_tagger = NewsMorphTagger(emb)
    syntax_parser = NewsSyntaxParser(emb)
    ner_tagger = NewsNERTagger(emb)

    wb = load_workbook('./src/blank1.xlsx')
    ws = wb.active
    cnt = 0
    for pdf_document in pdf_documents:
        doc = fitz.open(pdf_document)
        print("\n\n------------------\n")
        print("Исходный документ: ", doc)
        print("\nКоличество страниц: %i\n\n------------------\n" % doc.page_count)
        print(doc.metadata)
        for current_page in doc:
            ws.cell(row=4 + cnt, column=1, value=cnt + 1)
            page_text = get_page_text(current_page, doc)
            # print(page_text)
            # print(cnt + 1)
            date1 = re.search('от\s\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?\d(\s)?\d', page_text)
            if date1:
                date1 = re.search('\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?\d(\s)?\d', date1[0])
                ws.cell(row=4 + cnt, column=2, value=date1[0].replace(' ', ''))
                # print(date1[0].replace(' ', ''))
            code1 = re.search('8(\s)?6(\s)?0(\s)?1(\s)?0(\s)?/[\d\s]+/[\d\s]+', page_text)
            if code1:
                ws.cell(row=4 + cnt, column=3, value=code1[0].replace(' ', ''))
                # print(code1[0].replace(' ', ''))
            name = re.search('должника:[0-9a-zA-Zа-яА-ЯёЁ\s"\\\'.-]+,', page_text)
            if name:
                name = re.search(':[0-9a-zA-Zа-яА-ЯёЁ\s"\\\'.-]+', name[0])
                name = re.search('[0-9a-zA-Zа-яА-ЯёЁ"\\\'.-][0-9a-zA-Zа-яА-ЯёЁ\s"\\\'.-]+', name[0])
                name = name[0].replace("\n", " ").replace("- ", "")
                text = 'Почему у нас сегодня на работе нету ' + name
                d = Doc(text)
                d.segment(segmenter)
                d.tag_morph(morph_tagger)
                d.parse_syntax(syntax_parser)
                d.tag_ner(ner_tagger)
                name = ''
                for span in d.spans:
                    span.normalize(morph_vocab)
                    name += span.normal + ' '
                name = name.title()
                ws.cell(row=4 + cnt, column=7, value=name)
                # print(name)
            date2 = re.search(
                '\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?\d(\s)?\d\sго(-\n)?да\sро(-\n)?ж(-\n)?де(-\n)?ния',
                page_text)
            if date2:
                date2 = re.search('\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?\d(\s)?\d', date2[0])
                ws.cell(row=4 + cnt, column=6, value=date2[0].replace(' ', ''))
                # print(date2[0].replace(' ', ''))
            inn = re.search('ИНН\s[\d\s]+', page_text)
            if inn:
                inn = re.search('\d[\d\s]*', inn[0])
                ws.cell(row=4 + cnt, column=4, value=int(inn[0].replace(' ', '')))
                # print(inn[0].replace(' ', ''))
            code2 = re.search('[\d\s]+/[\d\s]+/[\d\s]+-(\s)?ИП', page_text)
            if code2:
                ws.cell(row=4 + cnt, column=5, value=code2[0].replace(' ', ''))
                # print(code2[0].replace(' ', ''))
            cnt += 1
        doc.close()

    if output_path[-5:] == '.xlsx':
        wb.save(output_path)
    else:
        wb.save(output_path + ".xlsx")

    return cnt


def doctype2():
    for pdf_document in pdf_documents:
        doc = fitz.open(pdf_document)
        print("\n\n------------------\n")
        print("Исходный документ: ", doc)
        print("\nКоличество страниц: %i\n\n------------------\n" % doc.page_count)
        print(doc.metadata)
        doc_text = ''
        for current_page in doc:
            doc_text += get_page_text(current_page, doc)
        doc.close()

        ish = re.search('[Ии]сх.[№0-9a-zA-Zа-яА-ЯёЁ "\\\'.«»-]+', doc_text)
        print(ish[0] if ish else '')

        name_uplim = re.search('Сведения\sо\sдолжнике:', doc_text)
        print(name_uplim.end() if name_uplim else '')
        name_downlim = re.search('Сведения\sо\sфинансовом\sуправляющем:', doc_text)
        print(name_downlim.start() if name_downlim else '')

        name = re.search('ФИО:\s[0-9a-zA-Zа-яА-ЯёЁ "\\\'-]+', doc_text[name_uplim:name_downlim])
        if name:
            name = re.search(':\s[0-9a-zA-Zа-яА-ЯёЁ "\\\'-]+', name[0])
            name = re.search('[0-9a-zA-Zа-яА-ЯёЁ"\\\'-][0-9a-zA-Zа-яА-ЯёЁ "\\\'-]+', name[0])
        print(name[0] if name else '')
        date = re.search('Дата рождения: \d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?\d(\s)?\d',
                         doc_text[name_uplim:name_downlim])
        if date:
            date = re.search('\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?.(\s)?\d(\s)?\d(\s)?\d(\s)?\d', date[0])
        print(date[0] if date else '')


def open_pdf_documents():
    global pdf_documents
    pdf_documents = filedialog.askopenfilenames(title="Выберите файлы формата PDF",
                                                filetypes=[("PDF files", "*.pdf")])
    entry_open.delete(0, END)
    entry_open.insert(0, pdf_documents)


def save_excel_file():
    global output_path
    output_path = filedialog.asksaveasfilename(title="Назовите файл")
    entry_save.delete(0, END)
    entry_save.insert(0, output_path)


def start_button():
    if pdf_documents == "" or output_path == "" or doctypes_var.get() == "":
        messagebox.showwarning("Предупреждение", "Вы не указали все данные")
    else:
        root.destroy()


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
# doctypes = ["Запрос от судебных приставов", "Запрос от финансового управляющего Коволенко"]
doctypes = ["Запрос от судебных приставов"]
pdf_documents = ''
output_path = ''
root = tk.Tk()
root.title("Docan")
root.geometry("500x300")
doctypes_var = tk.StringVar()
root.grid_rowconfigure(6, weight=1)
root.grid_columnconfigure(0, weight=1)

entry_open = ttk.Entry(width=60)
entry_open.grid(column=0, row=1, padx=5, pady=5, sticky="W")

open_button = Button(text="Открыть PDF файлы", command=open_pdf_documents, height=1, width=30)
open_button.grid(column=0, row=2, padx=5, pady=5, sticky="W")

entry_save = ttk.Entry(width=60)
entry_save.grid(column=0, row=3, padx=5, pady=5, sticky="W")

save_button = Button(text="Назвать и сохранить файл в ...", command=save_excel_file, height=1, width=30)
save_button.grid(column=0, row=4, padx=5, pady=5, sticky="W")

combobox = ttk.Combobox(textvariable=doctypes_var, values=doctypes, height=1, width=60)
combobox.grid(column=0, row=5, padx=5, pady=5, sticky="W")

start_button = Button(text="Начать", command=start_button, height=1, width=20)
start_button.grid(column=0, row=7, padx=5, pady=5, sticky="W")

root.mainloop()

doctype = doctypes_var.get()

if pdf_documents == "" or output_path == "" or doctypes_var.get() == "":
    exit(0)

if doctype == "Запрос от судебных приставов":
    doctype1()
elif doctype == "Запрос от финансового управляющего Коволенко":
    doctype2()
